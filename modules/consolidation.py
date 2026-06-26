"""
Consolidation module for exporting data to Excel and other formats.

Test with `pytest tests/test_EnvironmentData.py`
"""

import shutil
import datetime
import polars
import openpyxl
import logging
from typing import Dict, List, Callable

from modules import validation


def relocate(data: polars.DataFrame, columns: list) -> polars.DataFrame:
    """
    Helper function for relocating columns to the front of a DataFrame.

    Parameters
    ----------
    data : polars.DataFrame
        DataFrame to relocate columns in.

    columns : list[str]
        Columns to move to the front.

    Returns
    -------
    polars.DataFrame : DataFrame with columns relocated.
    """
    columns = [x for x in columns if x in data.columns]
    data = data[columns + [x for x in data.columns if x not in columns]]
    return data


def update_lookups(
    data_path: str,
    acceptable_range: Dict[str, List],
    logger: logging.Logger,
    error_callback: Callable[[str, bool], None] = None,
) -> None:
    """
    Build (or re-build) the lookup tables used for analytical queries using the consolidated parquet files: sensors, devices, and utcs.
    
    Parameters
    ----------
    data_path : str
        Path to the data directory containing parquet files.
    acceptable_range : Dict[str, List]
        Dictionary of reading types to acceptable ranges.
    logger : logging.Logger
        Logger instance for logging messages.
    error_callback : Callable[[str, bool], None], optional
        Callback function for error handling.
    """

    # DISABLED: Building info extraction from device names - data is too inconsistent
    # to reliably parse building information. Commenting out for now.
    #
    # # We need some manual fixes to reformat invalid names.
    # name_overrides = {
    #     "980D Unnamed Temp Sensor": "Temp Unnamed Unnamed_980D",
    #     "980D Unnamed Humid Sensor": "RH Unnamed Unnamed_980D",
    # }
    #
    # building_name_map = {
    #     "ESC": "Environmental Science Center",
    #     "YPM": "Yale Peabody Museum",
    #     "KGL": "Kline Geology Laboratory",
    #     "CSC": "Collection Studies Center (West Campus)",
    # }
    #
    # # Cardinal directions to look for in device names (ordered by length to match longer ones first)
    # cardinal_directions = ["NE", "NW", "SE", "SW", "N", "S", "E", "W"]
    # 
    # def extract_cardinal_direction(name: str) -> str:
    #     """Extract cardinal direction from device name patterns like __N__ or _NE_."""
    #     if name is None:
    #         return None
    #     for direction in cardinal_directions:
    #         # Check for patterns like __N__ or _NE_ (with underscores on both sides)
    #         if f"__{direction}__" in name or f"_{direction}_" in name:
    #             return direction
    #     return None
    #
    # def parse_device_name(device_name: str) -> dict:
    #     """Parse a device name to extract building info.
    #     
    #     Returns dict with: DeviceSerialFromName, BuildingID, Building, Room, CardinalDirection
    #     """
    #     if device_name is None:
    #         return {
    #             "DeviceSerialFromName": None,
    #             "BuildingID": None,
    #             "Building": None,
    #             "Room": None,
    #             "CardinalDirection": None,
    #         }
    #     
    #     # Try to extract cardinal direction from the device name pattern
    #     extracted_direction = extract_cardinal_direction(device_name)
    #     
    #     # Apply name overrides if needed
    #     name_to_parse = device_name
    #     if name_to_parse in name_overrides:
    #         name_to_parse = name_overrides[name_to_parse]
    #     
    #     # Fix names with problematic prefixes by removing everything up to and including the first space
    #     if "__" in name_to_parse and " " in name_to_parse:
    #         name_to_parse = name_to_parse[name_to_parse.index(" ") + 1:]
    #     
    #     # Fix names with " - no comm" suffix
    #     if " - no comm" in name_to_parse:
    #         name_to_parse = name_to_parse.replace(" - no comm", "")
    #     
    #     # Parse the name
    #     if "floator" in name_to_parse.lower():
    #         info = name_to_parse.strip().split("_")
    #     else:
    #         info = name_to_parse.strip().split(" ")
    #         if len(info) > 0 and "_" in info[-1]:
    #             info = info[0:-1] + info[-1].split("_")
    #     
    #     # Clean up empty strings from consecutive separators
    #     info = [part for part in info if part.strip() != ""]
    #     
    #     # Parse based on number of parts
    #     if len(info) == 5:
    #         cardinal_dir = extracted_direction if extracted_direction else info[3]
    #         building_id = info[1]
    #         return {
    #             "DeviceSerialFromName": info[4],
    #             "BuildingID": building_id,
    #             "Building": building_name_map.get(building_id, "Unknown"),
    #             "Room": info[2].replace("_", ""),
    #             "CardinalDirection": cardinal_dir,
    #         }
    #     elif len(info) == 4:
    #         building_id = info[1]
    #         return {
    #             "DeviceSerialFromName": info[3],
    #             "BuildingID": building_id,
    #             "Building": building_name_map.get(building_id, "Unknown"),
    #             "Room": info[2].replace("_", ""),
    #             "CardinalDirection": extracted_direction,
    #         }
    #     elif len(info) == 3:
    #         # Floaters
    #         building_id = info[1] if len(info) > 1 else None
    #         return {
    #             "DeviceSerialFromName": info[2] if len(info) > 2 else None,
    #             "BuildingID": "FLOATER",
    #             "Building": building_name_map.get(building_id, "Unknown") if building_id else "Unknown",
    #             "Room": "FLOATER",
    #             "CardinalDirection": extracted_direction,
    #         }
    #     else:
    #         # Malformed - return NAs with extracted cardinal direction
    #         return {
    #             "DeviceSerialFromName": None,
    #             "BuildingID": None,
    #             "Building": None,
    #             "Room": None,
    #             "CardinalDirection": extracted_direction,
    #         }

    # ============ SENSORS TABLE ============
    # Read unique sensor metadata from sensor_readings
    sensors_data = polars.read_parquet(
        f"{data_path}/sensor_readings.parquet",
        columns=["Source", "SensorID", "SensorName", "SensorType", "DeviceID"],
    ).unique()
    
    # Validate sensors
    sensors = validation.clean_validate_sensors(
        sensors=sensors_data, 
        acceptable_range=acceptable_range,
        logger=logger,
        step="update_lookups",
        error_callback=error_callback
    )
    
    # Sort and write sensors table
    sensors = sensors.sort(["Source", "SensorID"])
    sensors.write_parquet(f"{data_path}/sensors.parquet")

    # ============ DEVICES TABLE ============
    # Get unique devices from sensor_readings
    devices = polars.read_parquet(
        f"{data_path}/sensor_readings.parquet",
        columns=["Source", "DeviceID", "DeviceName"],
    ).unique()
    
    # DISABLED: Building info parsing - data is too inconsistent to reliably extract building info
    # device_records = []
    # for device in devices.iter_rows(named=True):
    #     parsed = parse_device_name(device["DeviceName"])
    #     
    #     # Log warning for unparseable device names
    #     if parsed["BuildingID"] is None:
    #         logger.warning(
    #             f"Could not parse building info from device name: {device['DeviceName']}. "
    #             f"Valid formats: 'BYCBA_0400410__N____' or similar patterns with building/room info."
    #         )
    #     
    #     device_records.append({
    #         "Source": device["Source"],
    #         "DeviceID": device["DeviceID"],
    #         "DeviceName": device["DeviceName"],
    #         **parsed,
    #     })
    # 
    # devices = polars.DataFrame(device_records, infer_schema_length=None)
    
    # Get sensor information for each device from the sensors table
    device_sensors_lookup = sensors.group_by("DeviceID").agg([
        polars.col("SensorID").unique().sort().str.join("|").alias("SensorIDs"),
        polars.col("SensorName").unique().sort().str.join("|").alias("SensorNames"),
        polars.col("SensorType").unique().sort().str.join("|").alias("SensorTypes")
    ])
    
    # Join sensor information to devices
    devices = devices.join(device_sensors_lookup, how="left", on="DeviceID")
    
    # Sort and reorder columns
    # DISABLED: Building-based sorting - using DeviceName instead
    # devices = devices.sort(["BuildingID", "Room", "DeviceName"])
    devices = devices.sort(["DeviceName"])
    devices = relocate(
        devices,
        # DISABLED: Building columns removed from relocate
        # ["Source", "DeviceID", "DeviceName", "SensorIDs", "SensorNames", "SensorTypes", "DeviceSerialFromName", "BuildingID", "Building", "Room", "CardinalDirection"],
        ["Source", "DeviceID", "DeviceName", "SensorIDs", "SensorNames", "SensorTypes"],
    )
    devices.write_parquet(f"{data_path}/devices.parquet")

    # UTC info.
    # Get all UTC timestamps from sensor_readings: both SensorReadingUTC and QueryUTC
    # sensor_readings is the primary source since all other data derives from it
    
    # Read both SensorReadingUTC and QueryUTC columns
    utc_data = polars.read_parquet(
        f"{data_path}/sensor_readings.parquet", 
        columns=["SensorReadingUTC", "QueryUTC"]
    )
    
    # Get SensorReadingUTC values (remove nulls)
    sensor_reading_utcs = (
        utc_data.filter(polars.col("SensorReadingUTC").is_not_null())
        ["SensorReadingUTC"]
        .unique()
        .to_list()
    )
    
    # Get QueryUTC values (remove nulls) 
    query_utcs = (
        utc_data.filter(polars.col("QueryUTC").is_not_null())
        ["QueryUTC"]
        .unique()
        .to_list()
    )
    
    # Combine all UTC timestamps and get unique values
    utc_timestamps = list(set(sensor_reading_utcs + query_utcs))

    utcs = polars.DataFrame(
        {
            "UTC": utc_timestamps,
            "datetime_utc": [
                datetime.datetime.fromtimestamp(x) for x in utc_timestamps
            ],
        }
    )

    # Convert to EST and round to seconds.
    utcs = utcs.with_columns(
        polars.col("datetime_utc")
        .dt.convert_time_zone("America/New_York")
        .dt.round("1s")
        .alias("datetime_est")
    )

    # Extract all the date parts.
    utcs = utcs.with_columns(
        polars.col("datetime_est").dt.date().alias("date"),
        polars.col("datetime_est").dt.time().alias("time"),
        polars.col("datetime_est").dt.year().alias("year"),
        polars.col("datetime_est").dt.month().alias("month"),
        (polars.col("datetime_est").dt.strftime("%A")).alias("day_of_week"),
        polars.col("datetime_est")
        .dt.weekday()
        .alias("day_of_week_monday1_sunday7"),
        polars.col("datetime_est").dt.hour().alias("hour_24"),
        (polars.col("datetime_est").dt.hour() % 12).alias("hour_12"),
        polars.col("datetime_est").dt.strftime("%p").alias("am_pm"),
    )

    # Write the table to a file.
    utcs.write_parquet(f"{data_path}/utcs.parquet")


def export_to_excel(data_path: str, home_directory: str, logger) -> None:
    """
    Export consolidated data to Excel using the template at templates/consolidated-data-sample-template.xlsx.
    Fills in sheets: sensors, devices, device_readings_daily, and device_readings_last1000.
    Saves the result to data/consolidated-data-sample.xlsx.
    
    The column order is determined by the template headers in row 1. Data columns are matched
    to template columns by name, so you can rearrange columns in the template and the code
    will follow. Columns in the data that are not in the template are ignored.

    Parameters
    ----------
    data_path : str
        Path to the data directory containing parquet files.
    home_directory : str
        Base directory for finding the template file.
    logger : logging.Logger
        Logger instance for logging messages.
    """
    template_path = f"{home_directory}/templates/consolidated-data-sample-template.xlsx"
    output_path = f"{data_path}/consolidated-data-sample.xlsx"
    
    # Copy template to output location
    shutil.copy(template_path, output_path)
    
    # Load the workbook
    wb = openpyxl.load_workbook(output_path)
    
    # Load the data
    sensors = polars.read_parquet(f"{data_path}/sensors.parquet")
    devices = polars.read_parquet(f"{data_path}/devices.parquet")
    utcs = polars.read_parquet(f"{data_path}/utcs.parquet")
    device_readings_daily = polars.read_parquet(f"{data_path}/device_readings_daily.parquet")
    device_readings = polars.read_parquet(f"{data_path}/device_readings.parquet")
    
    # Get first 500 and last 500 device readings and join with utcs for datetime_est
    sorted_readings = device_readings.sort("SensorReadingUTC")
    first_500 = sorted_readings.head(500)
    last_500 = sorted_readings.tail(500)
    device_readings_last1000 = (
        polars.concat([first_500, last_500], how="vertical")
        .join(utcs.select(["UTC", "datetime_est"]), left_on="SensorReadingUTC", right_on="UTC", how="left")
        .rename({"datetime_est": "reading_datetime_est"})
        .join(utcs.select(["UTC", "datetime_est"]), left_on="QueryUTC", right_on="UTC", how="left")
        .rename({"datetime_est": "query_datetime_est"})
    )
    
    # Map sheet names to data
    sheet_data = {
        "sensors": sensors,
        "devices": devices,
        "device_readings_daily": device_readings_daily,
        "device_readings_last1000": device_readings_last1000,
    }
    
    for sheet_name, df in sheet_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Read existing headers from row 1 of the template
            template_columns = []
            col_idx = 1
            while True:
                header = ws.cell(row=1, column=col_idx).value
                if header is None or header == "":
                    break
                template_columns.append(header)
                col_idx += 1
            
            # Capture number formats from the template's header row (row 1) for each column
            # The template has number formats defined on the header row
            column_number_formats = {}
            for col_idx in range(1, len(template_columns) + 1):
                template_cell = ws.cell(row=1, column=col_idx)
                column_number_formats[col_idx] = template_cell.number_format
            
            # If template has headers, use that order; otherwise use data columns
            if template_columns:
                # Filter to only columns that exist in both template and data
                columns_to_write = [col for col in template_columns if col in df.columns]
                missing_in_data = [col for col in template_columns if col not in df.columns]
                if missing_in_data:
                    logger.warning(f"Sheet '{sheet_name}': Template columns not in data: {missing_in_data}")
            else:
                # No template headers - write data columns and their headers
                columns_to_write = df.columns
                for col_idx, col_name in enumerate(columns_to_write, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data starting from row 2, following template column order
            for row_idx, row in enumerate(df.iter_rows(named=True), start=2):
                for col_idx, col_name in enumerate(columns_to_write, start=1):
                    value = row[col_name]
                    # Convert polars types to Python native types for Excel
                    if value is not None:
                        if hasattr(value, 'item'):
                            value = value.item()
                        # Strip timezone info from datetime values (Excel doesn't support timezones)
                        if hasattr(value, 'tzinfo') and value.tzinfo is not None:
                            value = value.replace(tzinfo=None)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Apply the template's number format to preserve formatting
                    if col_idx in column_number_formats and column_number_formats[col_idx] != 'General':
                        cell.number_format = column_number_formats[col_idx]
    
    # Save the workbook
    wb.save(output_path)
    wb.close()
    
    logger.info(f"Exported data to {output_path}")


def build_devices(
    data: polars.DataFrame,
    acceptable_range: Dict[str, List],
    error_callback: Callable[[str, bool], None] = None,
) -> polars.DataFrame:
    """
    Reformat the sensor data to create a DataFrame of devices.
    Each Device can have multiple sensors.
    In some cases it is easier to work with data indexed by Device with multiple types of readings (Temperature, Humidity)
    in the same row instead of Sensors which only have one type of reading per row.

    Parameters
    ----------
    data : polars.DataFrame
        DataFrame containing the sensor data.
    acceptable_range : Dict[str, List]
        Dictionary of acceptable ranges for each sensor type.
    error_callback : Callable[[str, bool], None], optional
        Callback function for error handling. Takes error message and raise_exception boolean.

    Returns
    -------
    polars.DataFrame: DataFrame containing the devices data.
    """

    # Get the data for each of the selected sensors.
    devices = None
    join_keys = ["Source", "DeviceID", "SensorReadingUTC", "QueryUTC"]
    data = data.filter(polars.col("DeviceID").is_null().not_())
    for reading in acceptable_range:
        if reading not in data.columns:
            continue
        idt = data.filter(polars.col(reading).is_null().not_()).select(
            join_keys + ["Historical", reading]
        )
        if idt.is_empty():
            continue
        if isinstance(devices, polars.DataFrame):
            devices = devices.join(
                idt,
                how="full",
                on=join_keys,
            )
            # Coalesce _right columns immediately to prevent accumulation across iterations
            for key_col in join_keys + ["Historical"]:
                right_cols = [c for c in devices.columns if c.startswith(key_col + "_right")]
                if right_cols:
                    devices = devices.with_columns(
                        polars.coalesce([key_col] + right_cols).alias(key_col)
                    ).drop(right_cols)
        else:
            devices = idt
        del idt, reading

    # If no reading types had data, create an empty devices DataFrame
    if devices is None:
        devices = data.select(join_keys + ["Historical"]).head(0)

    # If a device has multiple names, error out:
    device_names = (
        data.filter(polars.col("DeviceID").is_null().not_())
        .select(["DeviceID", "DeviceName"])
        .unique()
    )
    if device_names.shape[0] != device_names["DeviceName"].unique().shape[0]:
        # Identify the offending names so the log is actionable rather than opaque.
        dupes = (
            device_names.group_by("DeviceName")
            .agg(polars.col("DeviceID").n_unique().alias("n_ids"),
                 polars.col("DeviceID").unique().alias("ids"))
            .filter(polars.col("n_ids") > 1)
            .sort("n_ids", descending=True)
        )
        detail = "; ".join(
            f"{r['DeviceName']} -> {sorted(r['ids'])}" for r in dupes.iter_rows(named=True)
        )
        msg = (
            "DeviceName to DeviceID is not a 1-1 mapping "
            f"({dupes.shape[0]} name(s)): {detail}. "
            "Continuing; devices are keyed on DeviceID so the build is unaffected."
        )
        if error_callback:
            # Non-fatal: a duplicate sensor name (e.g. a device swap reusing a name)
            # is a recoverable data-quality condition. Log loudly and continue rather
            # than aborting consolidation, which would freeze the master write.
            error_callback(
                msg,
                False,
            )
        else:
            raise ValueError("DeviceName to DeviceID is not a 1-1 mapping.")

    # Attach the device name.
    devices = devices.join(device_names, how="left", on="DeviceID")

    # Get sensor information for each device
    device_sensors = (
        data.filter(polars.col("DeviceID").is_null().not_())
        .select(["DeviceID", "SensorID", "SensorName", "SensorType"])
        .unique()
        .group_by("DeviceID")
        .agg([
            polars.col("SensorID").sort().str.join("|").alias("Sensors"),
            polars.col("SensorName").sort().str.join("|").alias("SensorNames"), 
            polars.col("SensorType").sort().str.join("|").alias("SensorTypes")
        ])
    )
    
    # Join sensor information to devices
    devices = devices.join(device_sensors, how="left", on="DeviceID")

    # Rearrange columns (only include reading columns that exist in the data).
    reading_cols = [r for r in acceptable_range.keys() if r in devices.columns]
    devices = devices.select(
        ["Source", "DeviceID", "DeviceName", "Sensors", "SensorNames", "SensorTypes", "SensorReadingUTC", "QueryUTC", "Historical"]
        + reading_cols
    )

    # Return the data.
    return devices

def update_cubes(
    data_path: str,
    acceptable_range: Dict[str, List],
    logger: logging.Logger,
) -> None:
    """
    Build (or re-build) the cubed tables using the consolidated parquet files.
    Cubes contain aggregated measures by day and device/sensor to facilitate faster queries on smaller files.
    In the future, cubes can be modified to include different aggregation levels and dimensions.

    Parameters
    ----------
    data_path : str
        Path to the data directory containing parquet files.
    acceptable_range : Dict[str, List]
        Dictionary of acceptable ranges for each sensor type.
    logger : logging.Logger
        Logger instance for logging messages.
    """

    utcs = polars.read_parquet(f"{data_path}/utcs.parquet")

    sensor_readings = polars.read_parquet(
        f"{data_path}/sensor_readings.parquet"
    )
    # Only aggregate columns that exist in the data
    sumcols = [c for c in acceptable_range.keys() if c in sensor_readings.columns]
    sensor_readings = sensor_readings.filter(
        (polars.col("Historical") == True) | 
        ((polars.col("SensorReadingUTC") - polars.col("QueryUTC")).abs() < 60 * 5)
    )  # include historical data and recent readings within 5min window
    sensor_readings = sensor_readings.join(
        utcs[["UTC", "date"]],
        how="left",
        left_on="SensorReadingUTC",
        right_on="UTC",
    )

    sensor_readings_daily = (
        sensor_readings.group_by(["Source", "date", "SensorID"])
        .agg(
            [
                polars.len().alias("row_count"),
                polars.col(sumcols).sum().name.suffix("_sum"),
                polars.col(sumcols).min().name.suffix("_min"),
                polars.col(sumcols).max().name.suffix("_max"),
            ]
        )
        .sort(["Source", "date", "SensorID"])
    )

    sensor_readings_daily.write_parquet(
        f"{data_path}/sensor_readings_daily.parquet"
    )

    device_readings = polars.read_parquet(
        f"{data_path}/device_readings.parquet"
    )
    # Only aggregate columns that exist in device_readings
    device_sumcols = [c for c in acceptable_range.keys() if c in device_readings.columns]
    device_readings = device_readings.filter(
        (polars.col("Historical") == True) |
        ((polars.col("SensorReadingUTC") - polars.col("QueryUTC")).abs() < 60 * 5)
    )  # include historical data and recent readings within 5min window
    device_readings = device_readings.join(
        utcs[["UTC", "date"]],
        how="left",
        left_on="SensorReadingUTC",
        right_on="UTC",
    )

    device_readings_daily = (
        device_readings.group_by(["Source", "date", "DeviceID"])
        .agg(
            [
                polars.len().alias("row_count"),
                polars.col(device_sumcols).sum().name.suffix("_sum"),
                polars.col(device_sumcols).min().name.suffix("_min"),
                polars.col(device_sumcols).max().name.suffix("_max"),
            ]
        )
        .sort(["Source", "date", "DeviceID"])
    )

    device_readings_daily.write_parquet(
        f"{data_path}/device_readings_daily.parquet"
    )

    logger.info("Updated cubes: sensor_readings_daily and device_readings_daily")
